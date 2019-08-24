# ==================================================
# ==================================================
#   Working with Excel Files using openpyxl module
# ==================================================
# ==================================================

'''
# =================================================
# Create an xlsx workbook file using the openpyxl module
# Create a sheet in workbook
# Add top row as code1, code2, code3, code4, code5, code6
# Next add 10 rows of random data in each code1, ... code6 columns

import openpyxl as xl
import random

# Create a workbook
book = xl.Workbook()

# Creating a sheet
sheet = book.create_sheet('sheet1')

# Generate data for adding to sheet
topRow = [ "code"+ str(i) for i in range(1, 7)]

# Generate 10 rows
tenRows = []
for _ in range(10):
    tenRows.append([ random.randrange(10, 1000) for _ in range(6)])


# Append data to sheet
sheet.append(topRow)

for row in tenRows:
    sheet.append(row)

# Save the workbook
book.save('/home/hp/Downloads/Book1.xlsx')


# ==============================================
# Write a program to open existing sheet and change its values.

# Import modules
import openpyxl as xl
import random

# Open a workbook
book = xl.load_workbook('/home/hp/Downloads/Book1.xlsx')

# Open the sheet
sheet = book.get_sheet_by_name('sheet1')

# Replace data in the sheet
sheet['A1'] = 'Pass Code'

# Save the workbook
book.save('/home/hp/Downloads/Book1.xlsx')

'''
# ====================================================
# Write a program to open existing workbook and
# list all its sheets and active sheet in workbook

# Import modules
import openpyxl as xl

# Open the workbook
book = xl.load_workbook('/home/hp/Downloads/Book1.xlsx')

# List all the sheets in the book
print("All Sheets in workbook:", book.get_sheet_names())
print("Active Sheet:", book.get_active_sheet())

